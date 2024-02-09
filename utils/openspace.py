from .table import Table
from .table import Seat
from random import choice
from openpyxl import Workbook
from json import load

# Once again, our load_config. This is in case we change the amount of seats we have per table.
def load_config():
    with open("./utils/config.json", "r") as config_file:
        config = load(config_file)
    return config

config = load_config()
table_capacity = config["table_capacity"]
seats_table = []
table = Table(table_capacity, seats_table)
tables = []

# And our Openspace class
class Openspace:
    # We initialize a single variable, number_of_tables, which describes itself
    def __init__(self, number_of_tables, tables):
        self._number_of_tables = number_of_tables
        # We also create a List, called self._tables, using our weird list from main.py
        self._tables = tables
    
    # Not sure about this tbh, but it just returns the tables for now
    def __str__(self):
        return tables
    
    # Our organize function!
    def organize(self, names):
        """
        Function that will grab a list of names, and will organize them into seats and tables.

        :param names: List of names that we want to assign to seats.
        :return: Two variables, tables and remaining_names. Tables is all the tables after assigning seats to however many people we can. Remaining names is the people who we couldn't give seats to because there weren't enough.
        """
        # Clear existing tables before organizing new names
        tables = []
        # We create a copy of the names list, so we can later output any names we could seat, just in case there are any.
        remaining_names = names.copy()

        # We run a loop, we don't care about the variable in it.
        for _ in range(self._number_of_tables):
            # We assign a variable to our class Table
            table = Table(table_capacity, seats_table.copy())
            # We create an empty list, where we are going to assign people per table.
            assigned_names = []
            # We look at the remaining names, and queue them up
            for person in remaining_names:
                # And if there are still spots left to sit them on, we do so until it reaches its capacity.
                if len(table.seats) < table_capacity:
                    # If there are, we initialize a new Seat,
                    seat = Seat()
                    # Set the occupant to the person currently in queue
                    seat.set_occupant(person)
                    # And we assign them a seat at this table
                    table.assign_seat(seat)
                    # And so we append it to our table
                    assigned_names.append(person)
            # Once it reaches its capacity, we will append the table to tables.
            if table.seats:
                # Append a copy of the list of names, and then eventually restart our loop
                tables.append(table.seats.copy())  
            # We overwrite our remaining_names list, making sure to only include the names in it that are not in assigned_names
            remaining_names = [name for name in remaining_names if name not in assigned_names]
        # And we return our two variables, the tables, and the remaining names
        return tables, remaining_names

    def display_remaining_seats(self, tables, remaining_names):
        """
        Function that will display total, occupied and either leftover or missing seats.

        :param tables: our list of tables and their seats.
        :param remaining_names: the list of leftover people who couldn't get seats, can also be an empty list
        :return: Actually prints out the Total amount of Seats, as well as how many are occupied, and how many we have remaining.
        """
        # Total amount of seats is calculated by multiplying the number of tables by capacity
        total_seats = self._number_of_tables * table_capacity
        # Total amount of occupied seats is calculated by adding up the length of all the lists of table seats
        occupied_seats = sum(len(table) for table in tables)
        # And we get our remaining seats by a simple subtraction between total seats and occupied seats, and subtracting the length of remaining names
        remaining_seats = (total_seats - occupied_seats) - len(remaining_names)
        # Print total and occupied seats
        print(f"Total Seats: {total_seats}")
        print(f"Occupied Seats: {occupied_seats}")
        # If remaining seats is positive
        if remaining_seats >= 0:
            # Print leftover seats
            print(f"Leftover seats: {remaining_seats}")
        else:
            # Otherwise, print missing seats
            print(f"Missing seats: {abs(remaining_seats)}")
        
    def display(self, tables):
        """
        Function that will display total, occupied and either leftover or missing seats.

        :param tables: our list of tables and their seats.
        :return: Actually prints out the tables in a structured format.
        """
        # We iterate through table in the tables list
        for i, table in enumerate(tables, start=1):
            # Calls table by different numbers, depending on which iteration we are on
            print(f"Table {i}:")
            # Then we go through each table, and print out the names of people sitting on them
            for _, seat_name in enumerate(table, start=1):
                print(f"  {seat_name}")
            print()
    

    def store(self, filename, tables, remaining_names):
        """
        Function that will store our table information, as well as remaining names, into an excel file, using the workbook class from openpyxl

        :param filename: the filename it will save as
        :param tables: the list of tables
        :param remaining_names: the leftover people we couldn't seat.
        :return: Doesn't return anything, just saves it into an excel file.
        """
        # We create an excel workbook
        wb = Workbook()
        # We get the active worksheet in said notebook
        ws = wb.active

        #  Iterates through each table in the tables list and adds the table number and seat information to the worksheet, using tuples.
        for i, table in enumerate(tables, start=1):
            column_letter = chr(ord('A') + i - 1)
            ws[f'{column_letter}1'] = f'Table {i}'

            for j, seat in enumerate(table, start=2):
                ws[f'{column_letter}{j}'] = str(seat)

        # After adding all the tables, it adds the remaining names to the worksheet in a separate column. 
        #The column for remaining names is determined based on the number of tables.
        remaining_column = chr(ord('A') + len(tables))
        ws[f'{remaining_column}1'] = 'Remaining Names'

        for j, name in enumerate(remaining_names, start=2):
            ws[f'{remaining_column}{j}'] = name

        # Saves the workbook to the specified filename.
        wb.save(filename)

    # Some more properties
    @property
    def number_of_tables(self):
        return self._number_of_tables

    @property
    def tables(self):
        return self._tables
    
    """
    # WIP
    def add_occupant(self, name):
        for table in self.tables:
            if table.has_free_spot():
                table.assign_seat(name)
                return True
        return False
    
    # WIP
    def remove_occupant(self):
        # Randomly choose a table and seat
        table = choice(tables)
        seat = choice(table)

        old_occupant = seat.remove_occupant()

        return old_occupant
"""